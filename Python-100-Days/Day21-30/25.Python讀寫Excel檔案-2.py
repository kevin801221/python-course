#!/usr/bin/env python3
"""
從 25.Python读写Excel文件-2.md 提取的 Python 範例代碼
"""

# === 範例 1 ===
import datetime

import openpyxl

# 載入一個工作簿 ---> Workbook
wb = openpyxl.load_workbook('阿里巴巴2020年股票資料.xlsx')
# 獲取工作表的名字
print(wb.sheetnames)
# 獲取工作表 ---> Worksheet
sheet = wb.worksheets[0]
# 獲得單元格的範圍
print(sheet.dimensions)
# 獲得行數和列數
print(sheet.max_row, sheet.max_column)

# 獲取指定單元格的值
print(sheet.cell(3, 3).value)
print(sheet['C3'].value)
print(sheet['G255'].value)

# 獲取多個單元格（巢狀元組）
print(sheet['A2:C5'])

# 讀取所有單元格的資料
for row_ch in range(2, sheet.max_row + 1):
    for col_ch in 'ABCDEFG':
        value = sheet[f'{col_ch}{row_ch}'].value
        if type(value) == datetime.datetime:
            print(value.strftime('%Y年%m月%d日'), end='\t')
        elif type(value) == int:
            print(f'{value:<10d}', end='\t')
        elif type(value) == float:
            print(f'{value:.4f}', end='\t')
        else:
            print(value, end='\t')
    print()
# === 範例 2 ===
import random

import openpyxl

# 第一步：建立工作簿（Workbook）
wb = openpyxl.Workbook()

# 第二步：新增工作表（Worksheet）
sheet = wb.active
sheet.title = '期末成績'

titles = ('姓名', '語文', '數學', '英語')
for col_index, title in enumerate(titles):
    sheet.cell(1, col_index + 1, title)

names = ('關羽', '張飛', '趙雲', '馬超', '黃忠')
for row_index, name in enumerate(names):
    sheet.cell(row_index + 2, 1, name)
    for col_index in range(2, 5):
        sheet.cell(row_index + 2, col_index, random.randrange(50, 101))

# 第四步：儲存工作簿
wb.save('考試成績表.xlsx')
# === 範例 3 ===
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# 對齊方式
alignment = Alignment(horizontal='center', vertical='center')
# 邊框線條
side = Side(color='ff7f50', style='mediumDashed')

wb = openpyxl.load_workbook('考試成績表.xlsx')
sheet = wb.worksheets[0]

# 調整行高和列寬
sheet.row_dimensions[1].height = 30
sheet.column_dimensions['E'].width = 120

sheet['E1'] = '平均分'
# 設定字型
sheet.cell(1, 5).font = Font(size=18, bold=True, color='ff1493', name='華文楷體')
# 設定對齊方式
sheet.cell(1, 5).alignment = alignment
# 設定單元格邊框
sheet.cell(1, 5).border = Border(left=side, top=side, right=side, bottom=side)
for i in range(2, 7):
    # 公式計算每個學生的平均分
    sheet[f'E{i}'] = f'=average(B{i}:D{i})'
    sheet.cell(i, 5).font = Font(size=12, color='4169e1', italic=True)
    sheet.cell(i, 5).alignment = alignment

wb.save('考試成績表.xlsx')
# === 範例 4 ===
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

wb = Workbook(write_only=True)
sheet = wb.create_sheet()

rows = [
    ('類別', '銷售A組', '銷售B組'),
    ('手機', 40, 30),
    ('平板', 50, 60),
    ('筆記本', 80, 70),
    ('外圍裝置', 20, 10),
]

# 向表單中新增行
for row in rows:
    sheet.append(row)

# 建立圖表物件
chart = BarChart()
chart.type = 'col'
chart.style = 10
# 設定圖表的標題
chart.title = '銷售統計圖'
# 設定圖表縱軸的標題
chart.y_axis.title = '銷量'
# 設定圖表橫軸的標題
chart.x_axis.title = '商品類別'
# 設定資料的範圍
data = Reference(sheet, min_col=2, min_row=1, max_row=5, max_col=3)
# 設定分類的範圍
cats = Reference(sheet, min_col=1, min_row=2, max_row=5)
# 給圖表新增資料
chart.add_data(data, titles_from_data=True)
# 給圖表設定分類
chart.set_categories(cats)
chart.shape = 4
# 將圖表新增到表單指定的單元格中
sheet.add_chart(chart, 'A10')

wb.save('demo.xlsx')
